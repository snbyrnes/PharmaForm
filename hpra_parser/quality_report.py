"""Quality Summary Report generation for HPRA parser batch processing."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ProcessingResult:
    """Result of processing a single XML file."""
    filename: str
    status: str  # 'success', 'skipped', 'warning', 'failure'
    timestamp: datetime
    error_message: Optional[str] = None
    warning_message: Optional[str] = None
    records_processed: int = 0


@dataclass
class BatchMetrics:
    """Accumulated metrics for a batch processing run."""
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None
    results: List[ProcessingResult] = field(default_factory=list)
    
    @property
    def total_processed(self) -> int:
        """Count of successfully processed files."""
        return sum(1 for r in self.results if r.status == 'success')
    
    @property
    def total_skipped(self) -> int:
        """Count of skipped files."""
        return sum(1 for r in self.results if r.status == 'skipped')
    
    @property
    def total_warnings(self) -> int:
        """Count of files with warnings."""
        return sum(1 for r in self.results if r.status == 'warning')
    
    @property
    def total_failures(self) -> int:
        """Count of failed files."""
        return sum(1 for r in self.results if r.status == 'failure')
    
    @property
    def total_files(self) -> int:
        """Total number of files encountered."""
        return len(self.results)
    
    @property
    def total_records(self) -> int:
        """Total number of records processed across all files."""
        return sum(r.records_processed for r in self.results)
    
    def get_frequent_issues(self, top_n: int = 10) -> List[tuple[str, int]]:
        """Return the most frequent error/warning messages."""
        issue_counter = Counter()
        for result in self.results:
            if result.error_message:
                issue_counter[result.error_message] += 1
            if result.warning_message:
                issue_counter[result.warning_message] += 1
        return issue_counter.most_common(top_n)
    
    def add_result(self, result: ProcessingResult) -> None:
        """Add a processing result to the batch."""
        self.results.append(result)
    
    def finalize(self) -> None:
        """Mark the batch as complete."""
        self.end_time = datetime.now()


def generate_excel_report(metrics: BatchMetrics, output_path: Path) -> None:
    """Generate a human-readable Excel quality summary report."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required to generate Excel reports. "
            "Install it with: pip install openpyxl"
        )
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Summary Sheet
    summary_sheet = wb.create_sheet("Summary")
    _populate_summary_sheet(summary_sheet, metrics)
    
    # Detailed Results Sheet
    details_sheet = wb.create_sheet("Detailed Results")
    _populate_details_sheet(details_sheet, metrics)
    
    # Issues Sheet
    issues_sheet = wb.create_sheet("Frequent Issues")
    _populate_issues_sheet(issues_sheet, metrics)
    
    # Save the workbook
    wb.save(output_path)


def _populate_summary_sheet(ws, metrics: BatchMetrics) -> None:
    """Populate the Summary sheet with key metrics."""
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws.append(["HPRA Parser - Quality Summary Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    
    # Batch Information
    ws.append(["Batch Information"])
    ws["A3"].font = Font(bold=True, size=12)
    ws.append(["Start Time", metrics.start_time.strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["End Time", metrics.end_time.strftime("%Y-%m-%d %H:%M:%S") if metrics.end_time else "In Progress"])
    duration = (metrics.end_time - metrics.start_time).total_seconds() if metrics.end_time else 0
    ws.append(["Duration (seconds)", f"{duration:.2f}"])
    ws.append([])
    
    # Processing Counts
    ws.append(["Processing Counts"])
    ws["A8"].font = Font(bold=True, size=12)
    ws.append(["Metric", "Count"])
    ws["A9"].fill = header_fill
    ws["A9"].font = header_font
    ws["B9"].fill = header_fill
    ws["B9"].font = header_font
    
    ws.append(["Total Files", metrics.total_files])
    ws.append(["Successfully Processed", metrics.total_processed])
    ws.append(["Skipped", metrics.total_skipped])
    ws.append(["Warnings", metrics.total_warnings])
    ws.append(["Failures", metrics.total_failures])
    ws.append(["Total Records Processed", metrics.total_records])
    ws.append([])
    
    # Success Rate
    success_rate = (metrics.total_processed / metrics.total_files * 100) if metrics.total_files > 0 else 0
    ws.append(["Success Rate", f"{success_rate:.2f}%"])
    ws["A16"].font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _populate_details_sheet(ws, metrics: BatchMetrics) -> None:
    """Populate the Detailed Results sheet with per-file information."""
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Filename", "Status", "Timestamp", "Records", "Error/Warning Message"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for result in metrics.results:
        message = result.error_message or result.warning_message or ""
        ws.append([
            result.filename,
            result.status.upper(),
            result.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            result.records_processed,
            message
        ])
    
    # Apply conditional formatting for status
    for row_num in range(2, len(metrics.results) + 2):
        status_cell = ws.cell(row=row_num, column=2)
        status_value = status_cell.value
        
        if status_value == "SUCCESS":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        elif status_value == "FAILURE":
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
        elif status_value == "WARNING":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")
        elif status_value == "SKIPPED":
            status_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _populate_issues_sheet(ws, metrics: BatchMetrics) -> None:
    """Populate the Frequent Issues sheet."""
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws.append(["Most Frequent Issues"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    
    # Headers
    headers = ["Issue Description", "Occurrences"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    frequent_issues = metrics.get_frequent_issues(top_n=20)
    
    if not frequent_issues:
        ws.append(["No issues recorded", "0"])
    else:
        for issue, count in frequent_issues:
            ws.append([issue, count])
    
    # Column widths
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 15
    
    # Freeze header row
    ws.freeze_panes = "A4"


def generate_text_report(metrics: BatchMetrics) -> str:
    """Generate a plain text summary report (fallback if Excel is unavailable)."""
    lines = [
        "=" * 70,
        "HPRA Parser - Quality Summary Report",
        "=" * 70,
        "",
        "Batch Information:",
        f"  Start Time: {metrics.start_time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"  End Time: {metrics.end_time.strftime('%Y-%m-%d %H:%M:%S') if metrics.end_time else 'In Progress'}",
    ]
    
    if metrics.end_time:
        duration = (metrics.end_time - metrics.start_time).total_seconds()
        lines.append(f"  Duration: {duration:.2f} seconds")
    
    lines.extend([
        "",
        "Processing Counts:",
        f"  Total Files: {metrics.total_files}",
        f"  Successfully Processed: {metrics.total_processed}",
        f"  Skipped: {metrics.total_skipped}",
        f"  Warnings: {metrics.total_warnings}",
        f"  Failures: {metrics.total_failures}",
        f"  Total Records: {metrics.total_records}",
        "",
    ])
    
    success_rate = (metrics.total_processed / metrics.total_files * 100) if metrics.total_files > 0 else 0
    lines.append(f"Success Rate: {success_rate:.2f}%")
    lines.append("")
    
    frequent_issues = metrics.get_frequent_issues(top_n=10)
    if frequent_issues:
        lines.extend([
            "Most Frequent Issues:",
            ""
        ])
        for issue, count in frequent_issues:
            lines.append(f"  [{count}x] {issue}")
    
    lines.append("=" * 70)
    return "\n".join(lines)
